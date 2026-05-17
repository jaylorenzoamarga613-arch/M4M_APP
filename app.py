import streamlit as st
import datetime, io, uuid, json, os, base64, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# ─────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────
ADMIN_PASSWORD  = "bakyard2026"
ASSETS_DIR      = "assets"
EVENT_NAME      = "M4M: SIDLAKANG BALINGOAN"
EVENT_DATE      = "May 24, 2026"
EVENT_DATE_OBJ  = datetime.date(2026, 5, 24)
EVENT_LOCATION  = "Brgy. Dalihig, Balingoan, Misamis Oriental"
EVENT_ORGANIZER = "BAKYARD Events"
GCASH_NAME      = "Jorgil Amarga"
GCASH_NUMBER    = "0916 481 3822"
MAX_SLOTS       = 500   # change to your total slot limit

CATEGORIES  = {"7KM": 800, "15KM": 1200, "21KM": 1400}
SHIRT_SIZES = ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]
GENDERS     = ["Male", "Female", "Other / Prefer not to say"]
ASSET_KEYS  = ["shirt_7km","shirt_15km","shirt_21km","medal",
                "qr_gcash","qr_maya","banner","route_map"]
ASSET_LABELS = {
    "shirt_7km":"7KM Finisher Shirt","shirt_15km":"15KM Finisher Shirt",
    "shirt_21km":"21KM Finisher Shirt","medal":"Finisher Medal",
    "qr_gcash":"GCash QR Code","qr_maya":"Maya QR Code",
    "banner":"Event Banner","route_map":"Route Map",
}
WAIVERS = [
    ("Assumption of Risk",
     "I acknowledge that my participation involves inherent risks including slips, falls, "
     "uneven terrain, physical exertion, and weather conditions. Injuries may range from minor "
     "to severe or fatal. I voluntarily assume full responsibility for all risks.",
     "I have read, understand, and agree to assume all risks."),
    ("Fitness to Participate",
     "I represent that I am in good physical condition and have no medical condition that would "
     "prevent me from safely participating.",
     "I confirm I am medically and physically fit to participate."),
    ("Waiver and Release of Liability",
     "I hereby fully release, waive, and discharge the event organizers, directors, employees, "
     "volunteers, and sponsors from any and all claims arising out of any loss, damage, or injury "
     "sustained while participating.",
     "I agree to the waiver and release of liability."),
    ("Indemnification",
     "I agree to indemnify and hold harmless the event organizers from any costs, damages, "
     "lawsuits, or liabilities arising from my bodily injury, death, or property damage.",
     "I agree to the indemnification terms."),
    ("Medical Consent",
     "I authorize the event organizers and volunteers to secure emergency medical care or "
     "transportation if needed. I agree to assume all associated costs.",
     "I consent to emergency medical treatment and assume all associated costs."),
    ("Media Release",
     "I grant the event organizers the irrevocable right to photograph, record, and use my "
     "name, image, and likeness in any media for promotional purposes without compensation.",
     "I grant permission for the media release."),
    ("Digital Signature",
     "By checking this box, I agree that this digital submission acts as my legally binding "
     "signature for the M4M: SIDLAKANG | BALINGOAN event.",
     "I agree and submit my digital signature."),
]
SHEET_HEADERS = [
    "ID","Registered At","First Name","Last Name","Email","Phone","DOB","Age","Gender",
    "Address","Category","Fee","Shirt","Team","EC Name","EC Phone",
    "Medical","GCash Name","GCash No","Has Proof","Age Verify",
]

# ─────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="M4M: SIDLAKANG BALINGOAN",
    page_icon="🏔️",
    layout="wide",
    initial_sidebar_state="collapsed",
)
os.makedirs(ASSETS_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS — persistent storage
# ─────────────────────────────────────────────────────────────────
def get_sheet():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_dict = dict(st.secrets["gcp_service_account"])
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        sheet_name = st.secrets.get("SHEET_NAME", "Sidlakang Registrations")
        try:
            sh = client.open(sheet_name)
        except Exception:
            sh = client.create(sheet_name)
            sh.share(creds_dict["client_email"], perm_type="user", role="writer")
        for ws_name, cols in [("Registrations", SHEET_HEADERS),
                               ("Announcements", ["ID","Title","Body","Pinned","Date"])]:
            try:
                sh.worksheet(ws_name)
            except Exception:
                ws = sh.add_worksheet(ws_name, rows=2000, cols=30)
                ws.append_row(cols)
        return sh
    except Exception:
        return None

@st.cache_resource(ttl=20)
def load_registrations():
    sh = get_sheet()
    if not sh:
        return []
    try:
        rows = sh.worksheet("Registrations").get_all_records()
        return [{
            "id": str(r.get("ID","")),
            "registered_at": str(r.get("Registered At","")),
            "first_name": str(r.get("First Name","")),
            "last_name": str(r.get("Last Name","")),
            "email": str(r.get("Email","")),
            "contact": str(r.get("Phone","")),
            "dob": str(r.get("DOB","")),
            "age": str(r.get("Age","")),
            "gender": str(r.get("Gender","")),
            "address": str(r.get("Address","")),
            "category": str(r.get("Category","")),
            "fee": int(r.get("Fee",0) or 0),
            "shirt": str(r.get("Shirt","")),
            "team": str(r.get("Team","")),
            "emergency_name": str(r.get("EC Name","")),
            "emergency_phone": str(r.get("EC Phone","")),
            "medical": str(r.get("Medical","")),
            "pay_account_name": str(r.get("GCash Name","")),
            "pay_account_number": str(r.get("GCash No","")),
            "has_proof": str(r.get("Has Proof","")) == "Yes",
            "age_verify": str(r.get("Age Verify","")),
        } for r in rows]
    except Exception:
        return []

@st.cache_resource(ttl=20)
def load_announcements():
    sh = get_sheet()
    if not sh:
        return []
    try:
        rows = sh.worksheet("Announcements").get_all_records()
        return [{"id":str(r.get("ID","")), "title":str(r.get("Title","")),
                 "body":str(r.get("Body","")), "pinned":str(r.get("Pinned",""))=="True",
                 "date":str(r.get("Date",""))} for r in rows]
    except Exception:
        return []

def save_registration(reg):
    sh = get_sheet()
    if not sh:
        return False
    try:
        sh.worksheet("Registrations").append_row([
            reg.get("id",""), reg.get("registered_at",""),
            reg.get("first_name",""), reg.get("last_name",""),
            reg.get("email",""), reg.get("contact",""),
            reg.get("dob",""), str(reg.get("age","")),
            reg.get("gender",""), reg.get("address",""),
            reg.get("category",""), reg.get("fee",0),
            reg.get("shirt",""), reg.get("team",""),
            reg.get("emergency_name",""), reg.get("emergency_phone",""),
            reg.get("medical",""), reg.get("pay_account_name",""),
            reg.get("pay_account_number",""),
            "Yes" if reg.get("has_proof") else "No",
            reg.get("age_verify",""),
        ])
        load_registrations.clear()
        return True
    except Exception as e:
        st.warning(f"Could not save to sheet: {e}")
        return False

def save_announcement(ann):
    sh = get_sheet()
    if not sh: return False
    try:
        sh.worksheet("Announcements").append_row(
            [ann["id"],ann["title"],ann["body"],str(ann["pinned"]),ann["date"]])
        load_announcements.clear()
        return True
    except Exception:
        return False

def delete_announcement(ann_id):
    sh = get_sheet()
    if not sh: return
    try:
        ws   = sh.worksheet("Announcements")
        cell = ws.find(ann_id)
        if cell: ws.delete_rows(cell.row)
        load_announcements.clear()
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────
#  EMAIL
# ─────────────────────────────────────────────────────────────────
def send_email(reg):
    try:
        sender   = st.secrets.get("SENDER_EMAIL","")
        password = st.secrets.get("SENDER_PASSWORD","")
        if not sender or not password or not reg.get("email",""):
            return False
        name = f"{reg.get('first_name','')} {reg.get('last_name','')}".strip()
        rid  = reg.get("id","")
        cat  = reg.get("category","")
        fee  = reg.get("fee",0)

        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f5f0e8;margin:0;padding:20px}}
.wrap{{max-width:600px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)}}
.hdr{{background:linear-gradient(135deg,#050a07,#0f1f12);padding:32px;text-align:center}}
.hdr-title{{font-size:48px;font-weight:900;text-transform:uppercase;letter-spacing:-2px;color:#c9a84c;line-height:1}}
.hdr-sub{{font-size:14px;font-weight:700;letter-spacing:4px;text-transform:uppercase;color:#8a9a8a;margin-top:6px}}
.body{{padding:32px}}
.greet{{font-size:20px;font-weight:700;color:#0f1f12;margin-bottom:8px}}
.intro{{font-size:14px;color:#555;line-height:1.8;margin-bottom:24px}}
.badge{{display:inline-block;background:#0f1f12;color:#c9a84c;font-size:12px;font-weight:700;
  letter-spacing:3px;text-transform:uppercase;padding:8px 20px;border-radius:4px;margin-bottom:24px}}
.card{{background:#f7f3ec;border:1px solid #e0d8c8;border-radius:6px;padding:18px 22px;margin-bottom:20px}}
.card-title{{font-size:10px;font-weight:700;letter-spacing:3px;text-transform:uppercase;color:#c9a84c;margin-bottom:12px}}
.row{{display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #e8e2d4;font-size:13px}}
.row:last-child{{border:none}}
.rl{{color:#888}}.rv{{color:#1a1a1a;font-weight:600;text-align:right}}
.note{{background:#fff9ee;border:1px solid #f0e4b8;border-left:4px solid #c9a84c;
  border-radius:4px;padding:14px 18px;font-size:13px;color:#666;line-height:1.7;margin-bottom:20px}}
.ftr{{background:#0a0f0d;padding:20px;text-align:center}}
.ftr p{{font-size:11px;color:#8a9a8a;margin:3px 0}}
.ftr a{{color:#c9a84c;text-decoration:none}}
</style></head><body>
<div class="wrap">
<div class="hdr">
  <div style="font-size:10px;font-weight:700;letter-spacing:4px;text-transform:uppercase;color:#c9a84c;opacity:.8;margin-bottom:8px">Miles for Minds &bull; Inaugural Edition</div>
  <div class="hdr-title">SIDLAKANG</div>
  <div class="hdr-sub">Balingoan</div>
</div>
<div class="body">
  <div class="greet">Congratulations, {name}! 🏔️</div>
  <p class="intro">Your registration for <strong>M4M: SIDLAKANG BALINGOAN</strong> has been received and is pending payment verification. This email is your official registration confirmation.</p>
  <div class="badge">Registration ID: {rid}</div>
  <div class="card">
    <div class="card-title">Registration Details</div>
    <div class="row"><span class="rl">Full Name</span><span class="rv">{name}</span></div>
    <div class="row"><span class="rl">Category</span><span class="rv">{cat}</span></div>
    <div class="row"><span class="rl">Registration Fee</span><span class="rv">&#8369;{fee:,}</span></div>
    <div class="row"><span class="rl">T-Shirt Size</span><span class="rv">{reg.get('shirt','')}</span></div>
    <div class="row"><span class="rl">Team / Club</span><span class="rv">{reg.get('team','') or '—'}</span></div>
    <div class="row"><span class="rl">Payment Proof</span><span class="rv">{'Uploaded ✓' if reg.get('has_proof') else 'Pending'}</span></div>
  </div>
  <div class="card">
    <div class="card-title">Event Details</div>
    <div class="row"><span class="rl">Race Day</span><span class="rv">May 24, 2026</span></div>
    <div class="row"><span class="rl">Location</span><span class="rv">Brgy. Dalihig, Balingoan</span></div>
    <div class="row"><span class="rl">Organizer</span><span class="rv">BAKYARD Events</span></div>
  </div>
  <div class="note">⚠️ <strong>Note:</strong> Your slot is confirmed once GCash payment is verified by the organizer. Keep this email as your proof of registration. Questions? Message us on Facebook: <strong>Bakyard</strong>.</div>
</div>
<div class="ftr">
  <p><strong style="color:#c9a84c">M4M: SIDLAKANG BALINGOAN</strong></p>
  <p>May 24, 2026 &bull; Brgy. Dalihig, Balingoan, Misamis Oriental</p>
  <p>Presented by <a href="#">BAKYARD Events</a> &amp; LGU Balingoan</p>
  <p style="margin-top:10px;font-size:10px;opacity:.5">This is an automated email. Do not reply.</p>
</div>
</div></body></html>"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"✅ Registration Confirmed — {EVENT_NAME} [{rid}]"
        msg["From"]    = f"M4M SIDLAKANG <{sender}>"
        msg["To"]      = reg["email"]
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
            srv.login(sender, password)
            srv.sendmail(sender, reg["email"], msg.as_string())
        return True
    except Exception as e:
        st.warning(f"Email note: {e}")
        return False

# ─────────────────────────────────────────────────────────────────
#  ASSET HELPERS
# ─────────────────────────────────────────────────────────────────
def load_asset(key):
    path = os.path.join(ASSETS_DIR, f"{key}.png")
    try:
        return Image.open(path) if os.path.exists(path) else None
    except Exception:
        return None

def save_asset(key, img):
    img.save(os.path.join(ASSETS_DIR, f"{key}.png"))

# ─────────────────────────────────────────────────────────────────
#  EXCEL / CSV
# ─────────────────────────────────────────────────────────────────
def generate_excel(registrations):
    wb   = openpyxl.Workbook()
    hf   = Font(name="Arial", bold=True, color="C9A84C", size=11)
    hfill= PatternFill("solid", fgColor="0A0F0D")
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr  = Border(left=Side(style="thin",color="1C2A20"),
                  right=Side(style="thin",color="1C2A20"),
                  bottom=Side(style="thin",color="1C2A20"))
    ef   = PatternFill("solid", fgColor="111A14")
    df   = Font(name="Arial", size=10, color="E8E2D4")
    ws   = wb.active; ws.title="Registrations"; ws.sheet_properties.tabColor="C9A84C"
    headers=["#","Registered At","First Name","Last Name","Email","Phone","DOB","Age",
             "Gender","Address","Category","Fee (PHP)","Shirt","Team",
             "EC Name","EC Phone","Medical","GCash Name","GCash No","Proof","Age Verify"]
    widths=[4,20,13,13,26,16,12,5,10,26,10,10,8,16,18,16,20,18,13,8,14]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(row=1,column=c,value=h)
        cell.font=hf; cell.fill=hfill; cell.alignment=ctr
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=28
    for i,r in enumerate(registrations,2):
        fill=ef if i%2==0 else None
        vals=[i-1,r.get("registered_at",""),r.get("first_name",""),r.get("last_name",""),
              r.get("email",""),r.get("contact",""),r.get("dob",""),r.get("age",""),
              r.get("gender",""),r.get("address",""),r.get("category",""),r.get("fee",0),
              r.get("shirt",""),r.get("team",""),r.get("emergency_name",""),r.get("emergency_phone",""),
              r.get("medical",""),r.get("pay_account_name",""),r.get("pay_account_number",""),
              "Yes" if r.get("has_proof") else "No",r.get("age_verify","")]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=i,column=c,value=v); cell.font=df; cell.border=bdr
            if fill: cell.fill=fill
            if c==1: cell.alignment=ctr
            if c==12: cell.number_format='"₱"#,##0'
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:U{max(len(registrations)+1,2)}"
    ws2=wb.create_sheet("By Category"); ws2.sheet_properties.tabColor="C9A84C"
    for c,(h,w) in enumerate(zip(["Category","Fee","Registered","Revenue"],[14,12,12,16]),1):
        cell=ws2.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ctr
        ws2.column_dimensions[get_column_letter(c)].width=w
    for i,(cat,fee) in enumerate(CATEGORIES.items(),2):
        cnt=sum(1 for r in registrations if r.get("category")==cat)
        for c,v in enumerate([cat,fee,cnt,cnt*fee],1):
            cell=ws2.cell(row=i,column=c,value=v); cell.font=df; cell.border=bdr
            if i%2==0: cell.fill=ef
            if c in(2,4): cell.number_format='"₱"#,##0'
    tr=len(CATEGORIES)+2
    tfill=PatternFill("solid",fgColor="1E3A28")
    for c,v in enumerate(["TOTALS","",f"=SUM(C2:C{tr-1})",f"=SUM(D2:D{tr-1})"],1):
        cell=ws2.cell(row=tr,column=c,value=v)
        cell.font=Font(name="Arial",bold=True,size=10,color="C9A84C"); cell.fill=tfill
        if c==4: cell.number_format='"₱"#,##0'
    ws3=wb.create_sheet("Shirt Sizes"); ws3.sheet_properties.tabColor="C9A84C"
    for c,(h,w) in enumerate(zip(["Size","Count"],[10,10]),1):
        cell=ws3.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ctr
        ws3.column_dimensions[get_column_letter(c)].width=w
    for i,sz in enumerate(["XS","S","M","L","XL","XXL","XXXL"],2):
        cnt=sum(1 for r in registrations if r.get("shirt")==sz)
        for c,v in enumerate([sz,cnt],1):
            cell=ws3.cell(row=i,column=c,value=v); cell.font=df; cell.border=bdr
            if i%2==0: cell.fill=ef
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def generate_csv(registrations):
    import csv
    buf=io.StringIO(); w=csv.writer(buf)
    w.writerow(["#","Registered At","First Name","Last Name","Email","Phone","DOB","Age",
                "Gender","Address","Category","Fee","Shirt","Team","EC Name","EC Phone",
                "Medical","GCash Name","GCash No","Proof","Age Verify"])
    for i,r in enumerate(registrations,1):
        w.writerow([i,r.get("registered_at",""),r.get("first_name",""),r.get("last_name",""),
                    r.get("email",""),r.get("contact",""),r.get("dob",""),r.get("age",""),
                    r.get("gender",""),r.get("address",""),r.get("category",""),r.get("fee",0),
                    r.get("shirt",""),r.get("team",""),r.get("emergency_name",""),r.get("emergency_phone",""),
                    r.get("medical",""),r.get("pay_account_name",""),r.get("pay_account_number",""),
                    "Yes" if r.get("has_proof") else "No",r.get("age_verify","")])
    return buf.getvalue().encode()

# ─────────────────────────────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────────────────────────────
if "initialized" not in st.session_state:
    st.session_state.admin_logged_in = False
    st.session_state.step            = 1
    st.session_state.form            = {}
    st.session_state.proof_bytes     = None
    st.session_state.proof_name      = None
    st.session_state.proof_mime      = None
    st.session_state.success_id      = None
    st.session_state.success_reg     = None
    st.session_state.initialized     = True

# ─────────────────────────────────────────────────────────────────
#  LOAD LIVE DATA
# ─────────────────────────────────────────────────────────────────
regs = load_registrations()
anns = load_announcements()
total   = len(regs)
revenue = sum(r.get("fee",0) for r in regs)
today_n = sum(1 for r in regs if str(r.get("registered_at","")).startswith(str(datetime.date.today())))
slots_left = MAX_SLOTS - total
days_left  = max(0, (EVENT_DATE_OBJ - datetime.date.today()).days)

# ─────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@600;700;800;900&family=Barlow:wght@400;500;600&display=swap');
#MainMenu,footer,header{visibility:hidden}
.block-container{padding-top:0!important;padding-bottom:2rem}
[data-testid="stSidebar"]{display:none}
:root{--gold:#c9a84c;--gold2:#e8c96a;--gold3:#8a6820;--dark:#0a0f0d;--dark2:#111a14;
  --dark3:#1c2a20;--green:#2a5c3a;--cream:#f5f0e8;--text:#e8e2d4;--muted:#8a9a8a;
  --red:#e74c3c;--border:rgba(201,168,76,.2)}
.stApp{background:var(--dark)!important}
section[data-testid="stMain"]{background:var(--dark)}

/* HERO */
.hero{background:linear-gradient(135deg,#050a07,#0f1f12,#0c1a10);padding:32px 40px 24px;
  margin:-1rem -1rem 0;border-bottom:2px solid var(--gold3);position:relative;overflow:hidden}
.hero::before{content:'';position:absolute;inset:0;
  background-image:linear-gradient(rgba(201,168,76,.04) 1px,transparent 1px),
  linear-gradient(90deg,rgba(201,168,76,.04) 1px,transparent 1px);
  background-size:36px 36px;pointer-events:none}
.hero-inner{position:relative;z-index:1;display:flex;align-items:center;
  justify-content:space-between;flex-wrap:wrap;gap:20px}
.hero-eyebrow{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:700;
  letter-spacing:5px;text-transform:uppercase;color:var(--gold);opacity:.8;margin-bottom:6px}
.hero-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(44px,7vw,80px);
  font-weight:900;line-height:.88;text-transform:uppercase;letter-spacing:-2px;
  background:linear-gradient(180deg,#e8d88a,#c9a84c,#8a6820);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.hero-sub{font-family:'Barlow Condensed',sans-serif;font-size:18px;font-weight:700;
  letter-spacing:6px;text-transform:uppercase;color:var(--muted);margin:6px 0 14px}
.hero-pill{display:inline-block;font-family:'Barlow Condensed',sans-serif;font-size:11px;
  font-weight:700;letter-spacing:2px;text-transform:uppercase;border:1px solid var(--border);
  border-radius:2px;padding:4px 13px;color:var(--muted);margin:0 6px 6px 0}
.stat-hero{text-align:center;min-width:80px}
.stat-hero-n{font-family:'Barlow Condensed',sans-serif;font-size:38px;font-weight:900;
  color:var(--gold);line-height:1}
.stat-hero-l{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:var(--muted);margin-top:2px}
.stat-hero.danger .stat-hero-n{color:#e74c3c}
.stat-hero.warn .stat-hero-n{color:#f39c12}

/* COUNTDOWN BANNER */
.countdown{background:rgba(201,168,76,.08);border:1px solid rgba(201,168,76,.25);
  border-radius:4px;padding:14px 20px;margin-bottom:20px;display:flex;
  align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}
.cd-item{text-align:center;min-width:80px}
.cd-num{font-family:'Barlow Condensed',sans-serif;font-size:36px;font-weight:900;
  color:var(--gold);line-height:1}
.cd-lbl{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:var(--muted);margin-top:2px}
.slots-bar-wrap{flex:1;min-width:200px}
.slots-bar-label{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:700;
  letter-spacing:2px;text-transform:uppercase;color:var(--gold);margin-bottom:6px}
.slots-bar-bg{background:rgba(255,255,255,.08);border-radius:20px;height:8px;overflow:hidden}
.slots-bar-fill{height:100%;border-radius:20px;background:var(--gold);transition:width .5s}
.slots-bar-fill.warn{background:#f39c12}
.slots-bar-fill.danger{background:#e74c3c}
.slots-text{font-size:12px;color:var(--muted);margin-top:5px}

/* STEP BAR */
.stepbar{display:flex;position:relative;margin-bottom:32px}
.stepbar::before{content:'';position:absolute;top:16px;left:0;right:0;
  height:1px;background:var(--border);z-index:0}
.sitem{flex:1;display:flex;flex-direction:column;align-items:center;gap:7px;z-index:1}
.snum{font-family:'Barlow Condensed',sans-serif;width:32px;height:32px;border-radius:50%;
  display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:700;
  background:var(--dark3);border:1px solid var(--border);color:var(--muted)}
.snum.active{background:var(--gold);border-color:var(--gold);color:#000}
.snum.done{background:var(--green);border-color:var(--green);color:#fff}
.slbl{font-family:'Barlow Condensed',sans-serif;font-size:10px;font-weight:700;
  letter-spacing:2px;text-transform:uppercase;color:var(--muted)}
.slbl.active{color:var(--gold)}.slbl.done{color:var(--green)}

/* PAYMENT BOX */
.pay-box{background:rgba(201,168,76,.07);border:1px solid rgba(201,168,76,.28);
  border-radius:4px;padding:18px 22px;margin-bottom:18px}
.pay-title{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:700;
  letter-spacing:3px;text-transform:uppercase;color:var(--gold);margin-bottom:12px}
.fee-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:12px}
.fee-card{background:rgba(0,0,0,.25);border:1px solid var(--border);border-radius:3px;padding:12px;text-align:center}
.fee-dist{font-family:'Barlow Condensed',sans-serif;font-size:22px;font-weight:900;color:var(--gold)}
.fee-amt{font-size:14px;font-weight:600;color:var(--text);margin-top:2px}

/* WAIVER */
.wbox{background:rgba(0,0,0,.2);border:1px solid var(--border);border-radius:3px;
  padding:14px 18px;margin-bottom:10px}
.wbox-title{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:700;
  letter-spacing:3px;text-transform:uppercase;color:var(--gold);margin-bottom:6px}
.wbox-text{font-size:12px;color:var(--muted);line-height:1.7}

/* SUMMARY ROWS */
.srow{display:flex;justify-content:space-between;padding:6px 0;
  border-bottom:1px solid rgba(255,255,255,.06);font-size:13px;gap:10px}
.srow:last-child{border:none}
.srow-lbl{color:var(--muted);flex-shrink:0}
.srow-val{color:var(--text);font-weight:500;text-align:right}

/* ADMIN CARDS */
.acard{background:var(--dark3);border:1px solid var(--border);border-radius:4px;
  padding:16px;text-align:center;margin-bottom:10px}
.acard-n{font-family:'Barlow Condensed',sans-serif;font-size:34px;font-weight:900;color:var(--gold);line-height:1}
.acard-l{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:var(--muted);margin-top:3px}

/* ANNOUNCEMENTS */
.ann-card{background:var(--dark3);border:1px solid var(--border);border-left:3px solid var(--gold);
  border-radius:3px;padding:14px 18px;margin-bottom:12px}
.ann-title{font-family:'Barlow Condensed',sans-serif;font-size:16px;font-weight:700;
  letter-spacing:1px;color:var(--cream);margin-bottom:4px}
.ann-body{font-size:13px;color:var(--muted);line-height:1.7}
.ann-date{font-size:11px;color:var(--muted);opacity:.6;margin-top:6px}
.pub-ann{background:var(--dark2);border:1px solid var(--border);
  border-left:3px solid var(--gold3);border-radius:3px;padding:14px 18px;margin-bottom:10px}

/* STREAMLIT OVERRIDES */
.stApp *{color:var(--text)}
.stTabs [data-baseweb="tab-list"]{background:var(--dark2);border-bottom:2px solid var(--gold3);gap:0}
.stTabs [data-baseweb="tab"]{font-family:'Barlow Condensed',sans-serif!important;font-size:13px!important;
  font-weight:700!important;letter-spacing:2px!important;text-transform:uppercase!important;
  color:var(--muted)!important;border-bottom:3px solid transparent!important;padding:12px 18px!important}
.stTabs [aria-selected="true"]{color:var(--gold)!important;
  border-bottom:3px solid var(--gold)!important;background:rgba(201,168,76,.07)!important}
.stTabs [data-baseweb="tab-panel"]{padding-top:24px!important}
.stTextInput input,.stTextArea textarea,[data-baseweb="input"] input,[data-baseweb="textarea"] textarea{
  background:rgba(255,255,255,.04)!important;border:1px solid var(--border)!important;
  border-radius:3px!important;color:#000000!important;font-family:'Barlow',sans-serif!important}
label,[data-testid="stWidgetLabel"] p{font-family:'Barlow Condensed',sans-serif!important;
  font-size:11px!important;font-weight:700!important;letter-spacing:2px!important;
  text-transform:uppercase!important;color:rgba(201,168,76,.85)!important}
.stRadio label,.stCheckbox label{font-family:'Barlow',sans-serif!important;font-size:14px!important;
  font-weight:400!important;letter-spacing:0!important;text-transform:none!important;color:var(--text)!important}
.stSelectbox [data-baseweb="select"]{background:rgba(255,255,255,.04)!important;
  border:1px solid var(--border)!important;border-radius:3px!important}
.stSelectbox [data-baseweb="select"] span{color:#000000!important}
.stButton>button{font-family:'Barlow Condensed',sans-serif!important;font-size:14px!important;
  font-weight:700!important;letter-spacing:3px!important;text-transform:uppercase!important;
  background:var(--gold)!important;color:#000!important;border:none!important;
  border-radius:3px!important;padding:11px 28px!important}
.stButton>button:hover{background:var(--gold2)!important}
.stDownloadButton>button{font-family:'Barlow Condensed',sans-serif!important;font-size:13px!important;
  font-weight:700!important;letter-spacing:2px!important;text-transform:uppercase!important;
  background:var(--green)!important;color:var(--gold)!important;
  border:1px solid rgba(201,168,76,.3)!important;border-radius:3px!important}
div[data-testid="stFileUploader"]{border:1px dashed rgba(201,168,76,.35);border-radius:3px;
  background:rgba(255,255,255,.02);padding:6px}
.stSuccess{background:rgba(46,200,100,.1)!important;border:1px solid rgba(46,200,100,.25)!important}
.stError{background:rgba(231,76,60,.1)!important;border:1px solid rgba(231,76,60,.25)!important}
.stInfo{background:rgba(201,168,76,.08)!important;border:1px solid rgba(201,168,76,.2)!important}
p,.stMarkdown p{color:var(--text)!important;font-family:'Barlow',sans-serif!important}
h1,h2,h3{font-family:'Barlow Condensed',sans-serif!important;text-transform:uppercase!important;
  letter-spacing:-1px!important;color:var(--cream)!important}
div[data-testid="stImage"] img{border-radius:4px;border:1px solid var(--border)}
[data-testid="stNumberInput"] input{background:rgba(255,255,255,.04)!important;
  border:1px solid var(--border)!important;color:#000000!important}
[data-testid="stDateInput"] input{background:rgba(255,255,255,.04)!important;
  border:1px solid var(--border)!important;color:#000000!important}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────
def step_bar(current):
    labels=["Personal Info","Payment","Waiver","Confirm"]
    items="".join(
        f"<div class='sitem'><div class='snum {'active' if i==current else 'done' if i<current else ''}'>{i}</div>"
        f"<div class='slbl {'active' if i==current else 'done' if i<current else ''}'>{lbl}</div></div>"
        for i,lbl in enumerate(labels,1))
    st.markdown(f"<div class='stepbar'>{items}</div>",unsafe_allow_html=True)

def summary_box(rows):
    html="".join(f"<div class='srow'><span class='srow-lbl'>{a}</span><span class='srow-val'>{b}</span></div>" for a,b in rows)
    st.markdown(f"<div style='background:rgba(255,255,255,.04);border:1px solid var(--border);border-radius:4px;padding:18px 22px'>{html}</div>",unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
#  HERO
# ─────────────────────────────────────────────────────────────────
slot_pct = min(100, (total / MAX_SLOTS) * 100)
bar_cls  = "danger" if slots_left <= 20 else "warn" if slots_left <= 100 else ""

st.markdown(f"""
<div class='hero'>
  <div class='hero-inner'>
    <div>
      <div class='hero-eyebrow'>Miles for Minds &bull; Inaugural Edition</div>
      <div class='hero-title'>SIDLAKANG</div>
      <div class='hero-sub'>Balingoan</div>
      <div>
        <span class='hero-pill'>{EVENT_DATE}</span>
        <span class='hero-pill'>Brgy. Dalihig</span>
        <span class='hero-pill'>{EVENT_ORGANIZER}</span>
        <span class='hero-pill'>Balingoan, MisOr</span>
      </div>
    </div>
    <div style='display:flex;flex-direction:column;align-items:center;
      background:rgba(201,168,76,.08);border:1px solid rgba(201,168,76,.25);
      border-radius:6px;padding:18px 32px;text-align:center'>
      <div style='font-family:Barlow Condensed,sans-serif;font-size:72px;font-weight:900;
        line-height:1;background:linear-gradient(180deg,#e8d88a,#c9a84c);
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text'>
        {days_left}
      </div>
      <div style='font-family:Barlow Condensed,sans-serif;font-size:11px;font-weight:700;
        letter-spacing:4px;text-transform:uppercase;color:var(--muted);margin-top:4px'>
        Days to Race Day
      </div>
      <div style='font-family:Barlow Condensed,sans-serif;font-size:12px;font-weight:600;
        letter-spacing:2px;text-transform:uppercase;color:rgba(201,168,76,.5);margin-top:6px'>
        {EVENT_DATE}
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
#  ROUTE
# ─────────────────────────────────────────────────────────────────
params = st.query_params
is_admin = params.get("admin","") == "true"

if is_admin and not st.session_state.admin_logged_in:
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown("### 🔒 Admin Access")
    pwd=st.text_input("Password",type="password",placeholder="Enter admin password")
    if st.button("Login"):
        if pwd==ADMIN_PASSWORD:
            st.session_state.admin_logged_in=True; st.rerun()
        else:
            st.error("Incorrect password.")
    st.caption("For event staff only. [← Back to registration](./)")
    st.stop()

# ══════════════════════════════════════════════════════════════════
#  ADMIN
# ══════════════════════════════════════════════════════════════════
if is_admin and st.session_state.admin_logged_in:
    cl,cr=st.columns([1,6])
    with cl:
        if st.button("Logout"):
            st.session_state.admin_logged_in=False; st.query_params.clear(); st.rerun()
    with cr:
        if st.button("🔄 Refresh"):
            load_registrations.clear(); load_announcements.clear(); st.rerun()

    adm=st.tabs(["📊 Dashboard","👥 Registrants","🧾 Proofs","📣 Announcements","🖼 Materials","⬇ Export"])

    with adm[0]:
        st.markdown("### Dashboard")
        by7=sum(1 for r in regs if r.get("category")=="7KM")
        by15=sum(1 for r in regs if r.get("category")=="15KM")
        by21=sum(1 for r in regs if r.get("category")=="21KM")
        c1,c2,c3,c4,c5,c6=st.columns(6)
        for col,(n,l) in zip([c1,c2,c3,c4,c5,c6],[
            (total,"Registered"),(f"₱{revenue:,}","Revenue"),
            (today_n,"Today"),(f"{by7}/{by15}/{by21}","7K/15K/21K"),
            (slots_left,"Slots Left"),(f"{days_left}d","To Race Day")]):
            col.markdown(f"<div class='acard'><div class='acard-n'>{n}</div><div class='acard-l'>{l}</div></div>",unsafe_allow_html=True)
        if regs:
            import pandas as pd
            st.markdown("---")
            st.markdown("**By Category**")
            st.bar_chart(pd.DataFrame([{"Category":k,"Count":sum(1 for r in regs if r.get("category")==k)} for k in CATEGORIES]).set_index("Category"),color="#c9a84c")
            st.markdown("**Over Time**")
            dates={}
            for r in regs:
                d=str(r.get("registered_at",""))[:10]
                dates[d]=dates.get(d,0)+1
            if dates:
                st.line_chart(pd.DataFrame(sorted(dates.items()),columns=["Date","Count"]).set_index("Date"),color="#c9a84c")
            st.markdown("**Shirt Sizes**")
            st.bar_chart(pd.DataFrame([{"Size":s,"Count":sum(1 for r in regs if r.get("shirt")==s)} for s in ["XS","S","M","L","XL","XXL","XXXL"]]).set_index("Size"),color="#2a5c3a")

    with adm[1]:
        st.markdown("### All Registrants")
        if regs:
            import pandas as pd
            fc1,fc2,fc3=st.columns(3)
            with fc1: fcat=st.selectbox("Category",["All"]+list(CATEGORIES.keys()),key="fcat")
            with fc2: fsh=st.selectbox("Shirt",["All"]+["XS","S","M","L","XL","XXL","XXXL"],key="fsh")
            with fc3: fsrch=st.text_input("Search name/email",placeholder="Search...",key="fsrch")
            f2=regs
            if fcat!="All": f2=[r for r in f2 if r.get("category")==fcat]
            if fsh!="All":  f2=[r for r in f2 if r.get("shirt")==fsh]
            if fsrch:
                q=fsrch.lower()
                f2=[r for r in f2 if q in (r.get("first_name","")+r.get("last_name","")).lower() or q in r.get("email","").lower()]
            st.caption(f"Showing {len(f2)} of {total}")
            st.dataframe(pd.DataFrame([{"#":i+1,"Name":f"{r['first_name']} {r['last_name']}",
                "Category":r.get("category",""),"Shirt":r.get("shirt",""),
                "Gender":r.get("gender",""),"Email":r.get("email",""),
                "Contact":r.get("contact",""),"Fee":f"₱{r.get('fee',0):,}",
                "GCash":r.get("pay_account_number",""),
                "Proof":"✓" if r.get("has_proof") else "✗",
                "Team":r.get("team","") or "—","Medical":r.get("medical",""),
                "Registered":str(r.get("registered_at",""))[:16],
            } for i,r in enumerate(reversed(f2))]),use_container_width=True,hide_index=True)
        else:
            st.info("No registrations yet.")

    with adm[2]:
        st.markdown("### Payment Proof Viewer")
        rwp=[r for r in regs if r.get("has_proof")]
        if not rwp:
            st.info("No payment proofs uploaded yet.")
        else:
            names=[f"{r['first_name']} {r['last_name']} — {r.get('category','')} — {str(r.get('registered_at',''))[:10]}" for r in rwp]
            sel=st.selectbox("Select registrant",names)
            r=rwp[names.index(sel)]; rid=r.get("id","")
            st.markdown(f"**{r['first_name']} {r['last_name']}** · {r.get('category','')} · ₱{r.get('fee',0):,}")
            proof_path=os.path.join(ASSETS_DIR,f"proof_{rid}.png")
            if os.path.exists(proof_path):
                st.image(proof_path,caption="Payment proof",use_container_width=True)
                with open(proof_path,"rb") as pf:
                    st.download_button("Download",pf.read(),file_name=f"proof_{rid}.png",mime="image/png")
            else:
                st.warning("Proof not found on this server instance (file storage resets on redeploy).")

    with adm[3]:
        st.markdown("### Announcements")
        with st.form("ann_form"):
            at=st.text_input("Title *",placeholder="Race Day Reminder")
            ab=st.text_area("Message *",placeholder="Assembly at 4:00 AM...",height=100)
            ap=st.checkbox("Pin to top of public page")
            if st.form_submit_button("Post Announcement"):
                if at.strip() and ab.strip():
                    ann={"id":str(uuid.uuid4())[:8],"title":at.strip(),"body":ab.strip(),
                         "pinned":ap,"date":datetime.datetime.now().strftime("%B %d, %Y %H:%M")}
                    save_announcement(ann); st.success("Posted!"); st.rerun()
                else:
                    st.error("Title and message required.")
        st.markdown("---")
        anns_fresh=load_announcements()
        if not anns_fresh:
            st.info("No announcements yet.")
        for ann in anns_fresh:
            pin=" 📌" if ann.get("pinned") else ""
            st.markdown(f"<div class='ann-card'><div class='ann-title'>{ann['title']}{pin}</div><div class='ann-body'>{ann['body']}</div><div class='ann-date'>{ann['date']}</div></div>",unsafe_allow_html=True)
            if st.button("Delete",key=f"del_{ann['id']}"):
                delete_announcement(ann["id"]); st.rerun()

    with adm[4]:
        st.markdown("### Materials & QR Codes")
        for grp,keys in [("Race Materials",["shirt_7km","shirt_15km","shirt_21km","medal"]),
                          ("Payment QR Codes",["qr_gcash","qr_maya"]),
                          ("Event Assets",["banner","route_map"])]:
            st.markdown(f"**{grp}**")
            cols=st.columns(len(keys))
            for col,key in zip(cols,keys):
                with col:
                    img=load_asset(key)
                    if img: st.image(img,caption=ASSET_LABELS[key],use_container_width=True)
                    else: st.markdown(f"<div style='border:1px dashed rgba(201,168,76,.3);border-radius:3px;padding:20px;text-align:center;font-size:11px;text-transform:uppercase;letter-spacing:1px;color:var(--muted)'>{ASSET_LABELS[key]}<br>No image</div>",unsafe_allow_html=True)
                    up=st.file_uploader(f"{'Replace' if img else 'Upload'} {ASSET_LABELS[key]}",type=["jpg","jpeg","png","webp"],key=f"up_{key}")
                    if up: save_asset(key,Image.open(up)); st.success("Saved!"); st.rerun()
            st.markdown("<br>",unsafe_allow_html=True)

    with adm[5]:
        st.markdown("### Export Data")
        if not regs:
            st.info("No registrations to export yet.")
        else:
            e1,e2,e3=st.columns(3)
            with e1:
                st.markdown("**Excel (.xlsx)**")
                st.caption("3 sheets: registrations, by category, shirt sizes")
                st.download_button("⬇ Download Excel",generate_excel(regs),
                    file_name=f"sidlakang_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
            with e2:
                st.markdown("**CSV (Google Sheets)**")
                st.caption("Import directly into Google Sheets")
                st.download_button("⬇ Download CSV",generate_csv(regs),
                    file_name=f"sidlakang_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",use_container_width=True)
            with e3:
                st.markdown("**JSON Backup**")
                st.caption("Full data backup")
                st.download_button("⬇ Download JSON",
                    json.dumps({"registrations":regs,"exported_at":str(datetime.datetime.now())},indent=2,default=str).encode(),
                    file_name=f"sidlakang_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",use_container_width=True)
            st.markdown("---")
            st.markdown("**Import CSV into Google Sheets:**")
            st.markdown("1. Download CSV → 2. Open Google Sheets → 3. File → Import → Upload → Replace spreadsheet")
    st.stop()

# ══════════════════════════════════════════════════════════════════
#  PUBLIC PAGE
# ══════════════════════════════════════════════════════════════════
pinned=[a for a in anns if a.get("pinned")]
for ann in pinned:
    st.markdown(f"<div class='pub-ann'><div style='font-family:Barlow Condensed,sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;color:var(--gold);margin-bottom:4px'>📌 {ann['title']}</div><div style='font-size:13px;color:var(--muted);line-height:1.7'>{ann['body']}</div><div style='font-size:11px;color:var(--muted);opacity:.6;margin-top:5px'>{ann['date']}</div></div>",unsafe_allow_html=True)

pub_tabs=st.tabs(["📝  Register","📣  Updates"])

with pub_tabs[0]:

    # COUNTDOWN + SLOTS BANNER
    bar_color = "#e74c3c" if slots_left<=20 else "#f39c12" if slots_left<=100 else "#c9a84c"
    st.markdown(f"""
    <div class='countdown'>
      <div style='display:flex;gap:24px;flex-wrap:wrap'>
        <div class='cd-item'>
          <div class='cd-num'>{days_left}</div>
          <div class='cd-lbl'>Days to Race</div>
        </div>
        <div class='cd-item'>
          <div class='cd-num' style='color:{"#e74c3c" if slots_left<=20 else "#f39c12" if slots_left<=100 else "var(--gold)"}'>{slots_left}</div>
          <div class='cd-lbl'>Slots Left</div>
        </div>
        <div class='cd-item'>
          <div class='cd-num'>{total}</div>
          <div class='cd-lbl'>Registered</div>
        </div>
      </div>
      <div class='slots-bar-wrap'>
        <div class='slots-bar-label'>Slots Filled — {slot_pct:.0f}%</div>
        <div class='slots-bar-bg'>
          <div class='slots-bar-fill' style='width:{slot_pct:.1f}%;background:{bar_color}'></div>
        </div>
        <div class='slots-text'>{total} of {MAX_SLOTS} slots taken &nbsp;·&nbsp; {slots_left} remaining</div>
      </div>
    </div>
    """,unsafe_allow_html=True)

    # SUCCESS
    if st.session_state.success_id:
        r=st.session_state.success_reg or {}
        st.markdown("## 🏔️ You're In!")
        st.success(f"**{r.get('first_name','')} {r.get('last_name','')}** — registered for **{r.get('category','')}**. See you on the trail!")
        summary_box([
            ("Name",f"{r.get('first_name','')} {r.get('last_name','')}"),
            ("Registration ID",r.get("id","")),
            ("Category",r.get("category","")),
            ("Fee",f"₱{r.get('fee',0):,}"),
            ("T-Shirt",r.get("shirt","")),
            ("Race Day",EVENT_DATE),
            ("Location",EVENT_LOCATION),
        ])
        st.markdown("<br>",unsafe_allow_html=True)
        st.info("📧 A confirmation email has been sent to your email address.")
        st.info("📣 Follow **Bakyard** on Facebook for route maps and race-day updates.")
        if st.button("Register Another Runner"):
            st.session_state.success_id=None; st.session_state.success_reg=None
            st.session_state.step=1; st.session_state.form={}
            st.session_state.proof_bytes=None; st.session_state.proof_name=None
            load_registrations.clear(); st.rerun()
        st.stop()

    step=st.session_state.step
    step_bar(step)

    # STEP 1
    if step==1:
        st.caption("All fields marked * are required.")
        f=st.session_state.form or {}
        c1,c2=st.columns(2)
        with c1: fullname=st.text_input("Full Name *",value=f.get("fullname",""),placeholder="Juan Dela Cruz")
        with c2: email=st.text_input("Email Address *",value=f.get("email",""),placeholder="juan@email.com")
        c3,c4=st.columns(2)
        with c3: phone=st.text_input("Phone Number *",value=f.get("phone",""),placeholder="+63 912 345 6789")
        with c4:
            dv=f.get("dob")
            if isinstance(dv,str) and dv:
                try: dv=datetime.date.fromisoformat(dv)
                except: dv=None
            dob=st.date_input("Date of Birth *",value=dv or datetime.date(2000,1,1),
                              min_value=datetime.date(1930,1,1),max_value=datetime.date.today())
        c5,c6=st.columns(2)
        with c5:
            td=datetime.date.today()
            auto_age=td.year-dob.year-((td.month,td.day)<(dob.month,dob.day))
            age=st.number_input("Age *",min_value=1,max_value=100,value=int(f.get("age",auto_age) or auto_age))
        with c6:
            gender=st.radio("Gender *",GENDERS,index=GENDERS.index(f.get("gender",GENDERS[0])) if f.get("gender") in GENDERS else 0,horizontal=True)
        address=st.text_input("Home Address *",value=f.get("address",""),placeholder="Street, Barangay, City, Province")
        c7,c8=st.columns(2)
        with c7:
            cat_list=list(CATEGORIES.keys())
            cat_idx=cat_list.index(f.get("category")) if f.get("category") in cat_list else 0
            category=st.selectbox("Category *",cat_list,index=cat_idx,format_func=lambda k:f"{k} — ₱{CATEGORIES[k]:,}")
        with c8:
            sh_list=["XS","S","M","L","XL","XXL","XXXL"]
            sh_idx=sh_list.index(f.get("shirt")) if f.get("shirt") in sh_list else 2
            shirt=st.selectbox("T-Shirt Size (Unisex) *",sh_list,index=sh_idx)
        team=st.text_input("Team / Running Club",value=f.get("team",""),placeholder="Optional")
        c9,c10=st.columns(2)
        with c9: ecname=st.text_input("Emergency Contact Name *",value=f.get("ecname",""),placeholder="Maria Dela Cruz")
        with c10: ecphone=st.text_input("Emergency Contact Phone *",value=f.get("ecphone",""),placeholder="+63 912 345 6789")
        medical=st.text_area("Medical Conditions / Allergies *",value=f.get("medical",""),
                             placeholder="Type 'None' if not applicable",height=80)
        st.caption("Type 'None' if you have no conditions our medics need to know about.")
        st.divider()
        _,bc=st.columns([3,1])
        with bc:
            if st.button("Continue →",use_container_width=True,key="s1next"):
                errs=[]
                if not fullname.strip(): errs.append("Full name required.")
                if not email.strip() or "@" not in email: errs.append("Valid email required.")
                if not phone.strip(): errs.append("Phone required.")
                if not address.strip(): errs.append("Address required.")
                if not ecname.strip(): errs.append("Emergency contact name required.")
                if not ecphone.strip(): errs.append("Emergency contact phone required.")
                if not medical.strip(): errs.append("Medical field required (type 'None' if none).")
                if errs:
                    for e in errs: st.error(e)
                else:
                    st.session_state.form.update({"fullname":fullname.strip(),"email":email.strip(),
                        "phone":phone.strip(),"dob":str(dob),"age":age,"gender":gender,
                        "address":address.strip(),"category":category,"shirt":shirt,
                        "team":team.strip(),"ecname":ecname.strip(),"ecphone":ecphone.strip(),
                        "medical":medical.strip()})
                    st.session_state.step=2; st.rerun()

    # STEP 2
    elif step==2:
        st.caption("Send your fee via GCash then upload your screenshot.")
        f=st.session_state.form or {}
        cat=f.get("category","—"); fee=CATEGORIES.get(cat,0)
        st.markdown(f"""<div class='pay-box'><div class='pay-title'>GCash Payment Details</div>
          <div style='font-size:14px;margin-bottom:3px'><span style='color:var(--muted)'>Account Name:</span> &nbsp;<strong>{GCASH_NAME}</strong></div>
          <div style='font-size:14px;margin-bottom:12px'><span style='color:var(--muted)'>GCash Number:</span> &nbsp;<strong>{GCASH_NUMBER}</strong></div>
          <div class='fee-grid'>
            <div class='fee-card'><div class='fee-dist'>7KM</div><div class='fee-amt'>&#8369;800</div></div>
            <div class='fee-card'><div class='fee-dist'>15KM</div><div class='fee-amt'>&#8369;1,200</div></div>
            <div class='fee-card'><div class='fee-dist'>21KM</div><div class='fee-amt'>&#8369;1,400</div></div>
          </div></div>""",unsafe_allow_html=True)
        qg=load_asset("qr_gcash"); qm=load_asset("qr_maya")
        if qg or qm:
            q1,q2=st.columns(2)
            with q1:
                if qg: st.image(qg,caption="GCash QR",use_container_width=True)
            with q2:
                if qm: st.image(qm,caption="Maya QR",use_container_width=True)
        st.info(f"Your category: **{cat} — ₱{fee:,}**")
        p1,p2=st.columns(2)
        with p1: payname=st.text_input("GCash / Maya Account Name *",value=f.get("payname",""),placeholder="Full name on account")
        with p2: paynumber=st.text_input("GCash Mobile Number (11 digits) *",value=f.get("paynumber",""),placeholder="09XXXXXXXXX",max_chars=11)
        proof=st.file_uploader("Payment Screenshot / Proof *",type=["jpg","jpeg","png","pdf"])
        if proof:
            st.session_state.proof_bytes=proof.read(); st.session_state.proof_name=proof.name
            st.session_state.proof_mime=proof.type; st.success(f"✓ {proof.name} uploaded")
        elif st.session_state.proof_name:
            st.success(f"✓ {st.session_state.proof_name} (already uploaded)")
        st.divider()
        b1,b2=st.columns([1,3])
        with b1:
            if st.button("← Back",key="s2back"): st.session_state.step=1; st.rerun()
        with b2:
            _,bc2=st.columns([2,1])
            with bc2:
                if st.button("Continue →",use_container_width=True,key="s2next"):
                    errs=[]
                    if not payname.strip(): errs.append("Account name required.")
                    if not paynumber.strip() or len(paynumber)!=11: errs.append("Valid 11-digit GCash number required.")
                    if not st.session_state.proof_bytes: errs.append("Please upload your payment screenshot.")
                    if errs:
                        for e in errs: st.error(e)
                    else:
                        st.session_state.form.update({"payname":payname.strip(),"paynumber":paynumber.strip()})
                        st.session_state.step=3; st.rerun()

    # STEP 3
    elif step==3:
        st.caption("Read and check all boxes to proceed. This is legally binding.")
        all_checked=True
        for i,(title,text,confirm) in enumerate(WAIVERS,1):
            st.markdown(f"<div class='wbox'><div class='wbox-title'>{i}. {title}</div><div class='wbox-text'>{text}</div></div>",unsafe_allow_html=True)
            if not st.checkbox(confirm,key=f"w{i}"): all_checked=False
        st.markdown("---")
        st.markdown("**Age Verification**")
        age_verify=st.radio("Select the option that applies *",["adult","guardian"],
            format_func=lambda x:("I am 18 years of age or older — I voluntarily agree to this waiver." if x=="adult"
                                  else "I am the Parent / Legal Guardian of a participant under 18."),key="agev")
        st.divider()
        b1,b2=st.columns([1,3])
        with b1:
            if st.button("← Back",key="s3back"): st.session_state.step=2; st.rerun()
        with b2:
            _,bc3=st.columns([2,1])
            with bc3:
                if st.button("Continue →",use_container_width=True,key="s3next"):
                    if not all_checked: st.error("Please check all waiver boxes.")
                    else:
                        st.session_state.form["ageVerify"]=age_verify
                        st.session_state.step=4; st.rerun()

    # STEP 4
    elif step==4:
        st.caption("Review your details before final submission.")
        f=st.session_state.form or {}
        fee=CATEGORIES.get(f.get("category",""),0)
        summary_box([
            ("Full Name",f.get("fullname","—")),("Email",f.get("email","—")),
            ("Phone",f.get("phone","—")),("Date of Birth",str(f.get("dob","—"))),
            ("Age",str(f.get("age","—"))),("Gender",f.get("gender","—")),
            ("Address",f.get("address","—")),("Category",f.get("category","—")),
            ("Fee",f"₱{fee:,}"),("T-Shirt",f.get("shirt","—")),
            ("Team",f.get("team","—") or "—"),
            ("Emergency",f"{f.get('ecname','—')} · {f.get('ecphone','—')}"),
            ("Medical",f.get("medical","—")),
            ("GCash",f"{f.get('payname','—')} · {f.get('paynumber','—')}"),
            ("Age Verification","18+ (Self)" if f.get("ageVerify")=="adult" else "Parent/Guardian"),
            ("Payment Proof",f"✓ {st.session_state.proof_name}" if st.session_state.proof_name else "None"),
        ])
        st.markdown("<br>",unsafe_allow_html=True)
        st.info("By clicking **Submit Registration**, you confirm all information is accurate and you have agreed to all waiver terms.")
        b1,b2=st.columns([1,3])
        with b1:
            if st.button("← Back",key="s4back"): st.session_state.step=3; st.rerun()
        with b2:
            _,bc4=st.columns([2,1])
            with bc4:
                if st.button("Submit Registration",use_container_width=True,key="submit"):
                    parts=f.get("fullname","").strip().split()
                    first=parts[0] if parts else ""
                    last=" ".join(parts[1:]) if len(parts)>1 else ""
                    rid=str(uuid.uuid4())[:8].upper()
                    reg={"id":rid,"registered_at":datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                         "first_name":first,"last_name":last,"email":f.get("email",""),
                         "contact":f.get("phone",""),"dob":str(f.get("dob","")),"age":f.get("age",""),
                         "gender":f.get("gender",""),"address":f.get("address",""),
                         "category":f.get("category",""),"fee":CATEGORIES.get(f.get("category",""),0),
                         "shirt":f.get("shirt",""),"team":f.get("team",""),
                         "emergency_name":f.get("ecname",""),"emergency_phone":f.get("ecphone",""),
                         "medical":f.get("medical",""),"pay_account_name":f.get("payname",""),
                         "pay_account_number":f.get("paynumber",""),
                         "has_proof":bool(st.session_state.proof_bytes),
                         "age_verify":f.get("ageVerify","")}
                    save_registration(reg)
                    if st.session_state.proof_bytes:
                        try:
                            Image.open(io.BytesIO(st.session_state.proof_bytes)).save(os.path.join(ASSETS_DIR,f"proof_{rid}.png"))
                        except Exception:
                            with open(os.path.join(ASSETS_DIR,f"proof_{rid}.json"),"w") as pf:
                                json.dump({"name":st.session_state.proof_name,"mime":st.session_state.proof_mime,
                                           "data":base64.b64encode(st.session_state.proof_bytes).decode()},pf)
                    send_email(reg)
                    st.session_state.success_id=rid; st.session_state.success_reg=reg
                    st.session_state.form={}; st.session_state.step=1
                    st.session_state.proof_bytes=None; st.session_state.proof_name=None
                    st.rerun()

with pub_tabs[1]:
    st.markdown("### Updates & Announcements")
    if not anns:
        st.info("No announcements yet. Check back closer to race day!")
    else:
        for ann in anns:
            pin=" 📌" if ann.get("pinned") else ""
            st.markdown(f"<div class='pub-ann'><div style='font-family:Barlow Condensed,sans-serif;font-size:16px;font-weight:700;letter-spacing:1px;color:var(--gold);margin-bottom:5px'>{ann['title']}{pin}</div><div style='font-size:13px;color:var(--muted);line-height:1.7'>{ann['body']}</div><div style='font-size:11px;color:var(--muted);opacity:.5;margin-top:6px'>{ann['date']}</div></div>",unsafe_allow_html=True)
